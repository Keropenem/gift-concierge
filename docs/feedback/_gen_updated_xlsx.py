# -*- coding: utf-8 -*-
"""20260520サービス改善.xlsx をコピーして、光齋対応反映版を生成する。"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import shutil
import datetime
import sys

sys.stdout.reconfigure(encoding="utf-8")

src = "20260520サービス改善.xlsx"
dst = "20260520サービス改善_光齋対応反映_20260520.xlsx"

# 画像含め元xlsxを丸ごとコピー
shutil.copy(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb["改善内容"]

COLOR_DONE = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")  # 薄緑: 今回対応
COLOR_HOLD = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 薄黄: 保留(プロンプト)
COLOR_PREV = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")  # 薄灰: 以前完了
COLOR_HEAD = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")  # 薄青: ヘッダ

updates = [
    # (excel_row, L列ステータス, 色, G列に追記する文章)
    (2, "完了", COLOR_PREV, "（以前完了済み・今回変更なし）"),
    (3, "保留（プロンプト）", COLOR_HOLD,
     "【5/20】プロンプト改修担当領域（Web検索→実在URLの厳格採用強化）。コード側ノータッチ。"),
    (4, "完了", COLOR_PREV, "（以前完了済み・今回変更なし）"),
    (5, "完了", COLOR_DONE,
     "【5/20対応】提案メッセージ下に3段階リアクション（笑/普通/悲）+任意コメント欄を実装。"
     "feedbackテーブル新設、同一メッセージは上書き保存。コメントは書きたい人だけでOK（会議方針反映）。"
     "評価ボタンは絵文字を大きく+横にラベル併記して読みやすく改良。"),
    (6, "保留（プロンプト）", COLOR_HOLD,
     "【5/20】受け手データ・送り手プロフィールはコード側で渡済。"
     "LLMに「既に把握済みなのでStep冒頭で改めて触れない」挙動を守らせるのはプロンプト調整領域。"),
    (7, "完了", COLOR_PREV, "（以前完了済み・今回変更なし）"),
    (8, "保留（プロンプト）", COLOR_HOLD, "【5/20】#2と同系統。プロンプト改修担当。"),
    (9, "保留（プロンプト）", COLOR_HOLD,
     "【5/20】「実在しWebで購入可能なURL必須」をプロンプトで強化する範疇。コード側でDB絞り込みを入れる前段階。"),
    (10, "完了", COLOR_DONE,
     "【5/20対応】SP（タッチデバイス）ではEnter＝改行に変更。送信は送信ボタンタップに統一。"
     "PC側は従来通り（Enter送信 / Shift+Enterで改行）。"),
    (11, "完了", COLOR_DONE,
     "【5/20対応】textareaを overflow-y-auto + max-h-[200px] に変更。長文入力時も内部スクロールで全行確認可能。"),
    (12, "完了", COLOR_DONE,
     "【5/20対応】フォーム式モード /chat/form を新設（A/Bテスト用に対話モードと両立）。"
     "Step 1-3を一括フォーム入力→既存APIに整形メッセージで投入。"
     "ログイン済みなら過去のプロフィール・受け手情報を初期値に自動反映。"
     "トップに「まとめてフォームで入力する」リンク追加。"),
    (13, "保留（プロンプト）", COLOR_HOLD,
     "【5/20】LLM出力に画像URLを含めさせる、または別の画像APIで補完する設計が必要。プロンプト＋外部API判断領域。"),
    (14, "完了", COLOR_DONE,
     "【5/20対応】3画面構成で実装: /auth/forgot-password（メール送信）→ /auth/callback（セッション確立）"
     "→ /auth/reset-password（新パスワード設定）。ログイン画面に「パスワードを忘れた方」リンク追加。"
     "Supabase Auth resetPasswordForEmail使用。"
     "※デプロイ時、SupabaseのメールテンプレートでReset Passwordのリダイレクト URLを本番ドメインに合わせる作業が別途必要。"),
    (15, "完了", COLOR_DONE,
     "【5/20対応】トップに2つの入口ボタン追加: 「ギフトについて相談する」/「自分が欲しい物を相談する」。"
     "クリックで /chat?intent=gift|self に遷移→受け手ピッカー+目的別ヒントを表示。"
     "受け手をクリックすると「〇〇への贈り物を考えています。」を自動送信して即チャット開始（過去登録した相手情報も自動反映）。"),
    (16, "保留（プロンプト）", COLOR_HOLD,
     "【5/20】プロンプト側で「方向性の中間提案を出してユーザー確認を取る」フローを組む必要。辻村さん側のチューニング領域。"),
    (17, "完了", COLOR_DONE,
     "【5/20対応】チャットヘッダー下に Step 1-7 プログレスバーを追加"
     "（1.あなた / 2.相手 / 3.関係性 / 4.分析 / 5.物語 / 6.提案 / 7.演出）。"
     "往復数+「Step 6」「商品名」等のキーワード検出で現在位置を推定し点灯。"),
    (18, "完了", COLOR_DONE,
     "【5/20対応】保存形式を改修。新規セッション作成時に純粋なユーザー入力を profile_input._originalMessage に分離保存し、"
     "復元APIで messages[0] のテキストをこれで差し替え。旧セッション向けに「【...】」ブロック除去の正規表現フォールバックも実装。"
     "今後の履歴再開では黒バブルが見えない。"),
    (19, "保留（プロンプト）", COLOR_HOLD,
     "【5/20】プロンプト側で「3つの方向性を提示してユーザー選択を待つ」フローを組む範疇。辻村さん側のチューニング領域。"),
]


def set_cell(row, col_letter, value, fill=None):
    cell = ws[f"{col_letter}{row}"]
    cell.value = value
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(wrap_text=True, vertical="top")


for excel_row, status, fill, note in updates:
    # L列: 実施ステータス
    set_cell(excel_row, "L", status, fill)
    # G列: 備考に追記
    if note:
        current_g = ws[f"G{excel_row}"].value
        if current_g and str(current_g).strip():
            new_g = str(current_g).rstrip() + "\n\n---\n" + note
        else:
            new_g = note
        set_cell(excel_row, "G", new_g, fill)
    # 行全体を色付け
    for col_idx in range(1, 13):
        cl = get_column_letter(col_idx)
        ws[f"{cl}{excel_row}"].fill = fill

# 会議由来の追加項目を 新規行(20) として追加
new_row = 20
ws.cell(row=new_row, column=1, value=19)
ws.cell(row=new_row, column=2, value="WEB")
ws.cell(row=new_row, column=3, value="プロンプト設定（/admin/prompt）")
ws.cell(row=new_row, column=5,
        value="プロンプト変更履歴を残してチューニング作業を振り返れるようにしたい（5/20会議要望）")
ws.cell(row=new_row, column=6,
        value="保存時に本文＋任意メモを履歴テーブルに記録、画面下に履歴一覧（本文展開・エディタへ復元可能）")
ws.cell(row=new_row, column=7,
        value="【5/20対応】保存時に prompt_history テーブルに本文＋任意メモを自動記録。"
              "画面下部の履歴一覧から本文展開・エディタへ復元可能。"
              "辻村さんがチューニング時に「いつ・何を狙って変えたか」を後から振り返れる。")
ws.cell(row=new_row, column=8, value="朝香（会議発言）")
ws.cell(row=new_row, column=9, value=datetime.datetime(2026, 5, 20))
ws.cell(row=new_row, column=12, value="完了")
for col_idx in range(1, 13):
    cl = get_column_letter(col_idx)
    ws[f"{cl}{new_row}"].fill = COLOR_DONE
    ws[f"{cl}{new_row}"].alignment = Alignment(wrap_text=True, vertical="top")

# ヘッダ行
for col_idx in range(2, 13):
    cl = get_column_letter(col_idx)
    cell = ws[f"{cl}1"]
    if cell.value:
        cell.fill = COLOR_HEAD
        cell.font = Font(bold=True)

# 列幅・行高
column_widths = {
    "A": 8, "B": 6, "C": 14, "D": 22, "E": 30, "F": 30,
    "G": 55, "H": 10, "I": 12, "J": 14, "K": 14, "L": 18,
}
for col, w in column_widths.items():
    ws.column_dimensions[col].width = w
for row_idx in range(2, 21):
    if ws.row_dimensions[row_idx].height is None or (ws.row_dimensions[row_idx].height or 0) < 90:
        ws.row_dimensions[row_idx].height = 90

# 凡例シートを先頭に追加
if "凡例" in wb.sheetnames:
    del wb["凡例"]
legend = wb.create_sheet("凡例", 0)
legend["A1"] = "凡例（実施ステータス色分け） / 光齋対応反映版 2026-05-20"
legend["A1"].font = Font(bold=True, size=14)

rows = [
    ("色", "ステータス", "意味"),
    ("緑（薄）", "完了【5/20対応】", "今回のコード修正で対応完了。デプロイ後に動作確認をお願いします"),
    ("灰（薄）", "完了（以前完了済み）", "今回より前に完了済み・今回変更なし"),
    ("黄（薄）", "保留（プロンプト）", "中で動いている Gemini API のプロンプト改修が必要。辻村さん側の /admin/prompt チューニング待ち"),
    ("青（薄）", "—", "ヘッダ行"),
]
for r, row in enumerate(rows, start=3):
    for c, val in enumerate(row, start=1):
        legend.cell(row=r, column=c, value=val)
        if r == 3:
            legend.cell(row=r, column=c).font = Font(bold=True)
            legend.cell(row=r, column=c).fill = COLOR_HEAD
        else:
            cell = legend.cell(row=r, column=c)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if c == 1:
                if "緑" in val:
                    cell.fill = COLOR_DONE
                elif "灰" in val:
                    cell.fill = COLOR_PREV
                elif "黄" in val:
                    cell.fill = COLOR_HOLD
                elif "青" in val:
                    cell.fill = COLOR_HEAD

legend.column_dimensions["A"].width = 12
legend.column_dimensions["B"].width = 22
legend.column_dimensions["C"].width = 80

legend["A9"] = "今回の対応サマリー（光齋 / 2026-05-20）"
legend["A9"].font = Font(bold=True, size=12)
legend["A10"] = "[完了/今回対応] 8件: #4, #9, #10, #11, #13, #14, #16, #17, + 会議由来の#19（プロンプト変更履歴管理）"
legend["A11"] = "[完了/以前] 3件: #1, #3, #6"
legend["A12"] = "[保留/プロンプト] 7件: #2, #5, #7, #8, #12, #15, #18 ← 辻村さん側のチューニング待ち"
legend["A14"] = "デプロイ前のユーザー作業:"
legend["A14"].font = Font(bold=True)
legend["A15"] = "1. Supabaseで 00008_add_feedback.sql, 00009_add_prompt_history.sql を適用"
legend["A16"] = "2. Supabase AuthのメールテンプレートでReset Passwordのリダイレクト URLを本番ドメインに設定"
for r in range(9, 17):
    legend.row_dimensions[r].height = 22

wb.save(dst)
print("作成完了:", dst)
